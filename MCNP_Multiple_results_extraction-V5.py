# -*- coding: utf-8 -*-
"""
Created on Mon Dec 19 09:26:59 2016

@author: Magnus Rannala

Comments: Wed Jan 25 11:35:00 2017

At present an exception is created which causes script to abort if an output file
that has not been completed is encountered.


"""

import pandas as pd
import numpy as np
import openpyxl as pxl
from openpyxl import formatting, styles
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime
import os
import sys
import tkinter as tk
import tkinter.filedialog

#==============================================================================
#                               Subroutines
#
#==============================================================================

def stripn(self):
    # this will remove the character '\n' for array elements
    for k in range(0, len(self)):
        self[k] = self[k].rstrip('\n')
    return self
    
def find_Nvalues(self, count):
    
    Ndata = 0
    # count lines until next blank
    # NB m will remain unaltered globally

    while len(self[count]) >= 3:
        if self[count] == '1tally fluctuation charts                              \n' or self[count] == '1mesh-based weight window generator                                                      print table 190\n' :
            count +=1
            break
        Ndata += 1
        count += 1
    return Ndata-1
    
def Remove_spaces(self):
    
    # This will remove all spaces from each element.
    for i in range(0,len(self)):
        self[i] = self[i].replace(" ","")
        
    return self

def Add_to_Tally(self, i, NPS, Ntal, tallynumbers, upr, lwr):
	# adds the tally No. NPS and the result for each of the tally columns,

	# imports just numbers of interest
	self = self[ lwr : upr ]
	
	# add NPS
	self = [NPS] + self
	
	# add tally number
	self = [tallynumbers[Ntal]] + self
	
	return self

#==============================================================================
#                                    Code
#
#==============================================================================

dataframes = []

# Specify file type
filetype = ".o"

# list of files to cycle over
filelist = []

# Get current working directory (cwd)
cwd = os.getcwd()

# add all files andding in filetype to filelist
for file in os.listdir(cwd):
    if file.endswith(filetype):
        filelist.append(file)

#==============================================================================
#                                    Extraction
#
#==============================================================================
       
# Extraction for file in filelist
for files in range(len(filelist)):
    
    filename = filelist[files]    
    
    afile = open(filename, 'r')
    
    running = True
    datarray = []
    ResultsArray = []
    ResultsArray.append(['Tally','NPS','Mean','Error','VoV','Slope','FoM'])
    i = 0
    
    # Append each line to an array element
    for line in afile:
        datarray.append(line)
        
    # define iterable for main text
    iterable = iter(range(0, len(datarray)-1, 1))
    
    # search for tally block
    while running == True:
        
        if datarray[i] == "1tally fluctuation charts                              \n":
            running = False
        i += 1
    
    # define iterable
    iterable = iter(range(i, len(datarray)-1, 1))
    
    n = i
    running =True
    
    # step line by line until end
    while running == True:
    
        # skip blank line
        if len(datarray[n]) < 3 and running == True:
    
            n += 1
    
            # skip blank line
            if len(datarray[n]) < 3 and running == True:
                n += 1
            
            elif datarray[n].split()[0]  == '***********************************************************************************************************************':
    
                running = False
    
        # check if the first word of each like 'tally'
        if datarray[n].split()[0] == 'tally' and running == True:   
    
            # strip numbers of tallies
            tallynumbers = datarray[n].split("tally")
        
            #strips blank first element
            tallynumbers = tallynumbers[1:]
        
            # Strip '\n'
            tallynumbers = stripn(tallynumbers)
        
            # Strip spaces
            tallynumbers = Remove_spaces(tallynumbers)          
    
            # find the number of readings presented by MCNP 
            Ndata = find_Nvalues(datarray, n)
        
            # if more than zero tallies present            
            if len(tallynumbers) > 0: 
    
    
                 # Takes line into array 
                 holdarray =  datarray[n + Ndata].split()                                                             
         
                 # Takes NPS value
                 NPS = holdarray[0]
    
        		# Add tally values to Results Array
                 ResultsArray.append(Add_to_Tally(holdarray, n, NPS, 0, tallynumbers, 6, 1 ))
                            
            # if more than ONE tally is present
            if len(tallynumbers) > 1:
    
                # Add tally values to Results Array
                ResultsArray.append(Add_to_Tally(holdarray, n, NPS, 1, tallynumbers, 11, 6 )) 
            
            # if more than TWO tallies are present
            if len(tallynumbers) > 2:
    
                # Add tally values to Results Array
                ResultsArray.append(Add_to_Tally(holdarray, n, NPS, 2, tallynumbers, 16, 11 )) 
    
        n += 1
    
        # Allows for when a weight window is used (line following tallies)
        if datarray[n] == '1mesh-based weight window generator                                                      print table 190\n':
            
            running = False         

    # Create dataframe for this file
    df = pd.DataFrame(ResultsArray,)
    
    # Sets the top row as the column values
    df.columns = df.iloc[0]
    # Sets the column 'Tally' as the index
    # df = df.set_index(['Tally'])
    # Removes the top row of the df
    df = df.ix[1:]
    
    # Add dataframe to master list
    dataframes.append(df)
    
##=============================================================================

RESULT = pd.concat((dataframes), keys=filelist)
     
# Creates new Excel with filename:
xlfilename = "Results--" + os.getcwd().rsplit("\\")[-1] + "_" + datetime.datetime.strftime(datetime.datetime.now(), '%H-%M-%S') + ".xlsx"

# opens excel
writer = pd.ExcelWriter(xlfilename)

# Sort RESULT alphabetical order
RESULT = RESULT.sort_index(ascending=True)

# Make numeric (float)
RESULT[list(RESULT.columns.values)] = RESULT[list(RESULT.columns.values)].astype(float)

print RESULT.dtypes 

# Write Dataframe to excel 
RESULT.to_excel(writer, "Sheet 1", merge_cells=False)

# Select worksheet
sheet = writer.sheets.get("Sheet 1")

#### FORMATS
red_font = styles.Font(size=14, bold=True, color=red_color_font)
red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')

# 10 STATISTICAL TEST FORMATTING
sheet.conditional_formatting.add('C2:I999', FormulaRule(formula=['$F2 > 0.05'], fill=red_fill, font=red_font))
sheet.conditional_formatting.add('C2:I999', FormulaRule(formula=['$G2 > 0.10'], fill=red_fill, font=red_font))
sheet.conditional_formatting.add('C2:I999', FormulaRule(formula=['0.00 < $H2 < 3.00'], fill=red_fill, font=red_font))
# sheet.conditional_formatting.add('G2:G999', CellIsRule(operator='greaterThan', formula=['.01'], fill=red_fill, font=red_font))
# sheet.conditional_formatting.add('H2:H999', CellIsRule(operator='between', formula=['0.0','3.0'], fill=red_fill, font=red_font))

# Save new excel
writer.save()

      
        